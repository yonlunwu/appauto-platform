from pathlib import Path
from typing import Any, Dict, List
import math
import os

from fastapi import APIRouter, HTTPException, status, Depends, Query
from fastapi.responses import FileResponse
from openpyxl import load_workbook

from llm_perf_platform.api.auth import get_current_user
from llm_perf_platform.api.schemas import (
    ArchiveRequest,
    ArchiveResponse,
    CancelTaskResponse,
    DeleteTaskResponse,
    DeployAMaaSRequest,
    DeployFTRequest,
    DeployResponse,
    HardwareInfoCollectRequest,
    HardwareInfoCollectResponse,
    RetryTaskResponse,
    TaskDetailResponse,
    TaskListResponse,
    TaskLogsResponse,
    TaskSummary,
    TestRunRequest,
    TestRunResponse,
    TestPerfViaAMaaSSkipLaunch,
    TestPerfViaAmaaSWithLaunch,
    TestPerfViaFTSkipLaunch,
    TestPerfViaFTWithLaunch,
    TestEvalViaAMaaSSkipLaunch,
    TestEvalViaAMaaSWithLaunch,
    TestEvalViaFTSkipLaunch,
    TestEvalViaFTWithLaunch,
    TestEvalResponse,
    # 导入新的 Payload 模型
    PerfTestPayload,
    EvalTestPayload,
    EnvDeployPayload,
    HardwareInfoPayload,
    LegacyTestPayload,
)
from llm_perf_platform.executor.base_executor import TaskType
from llm_perf_platform.executor.logger import LOG_DIR
from llm_perf_platform.models.db import get_session
from llm_perf_platform.models.user_account import UserAccount
from llm_perf_platform.services.task_service import TaskService
from llm_perf_platform.storage.results import ResultStorage
from llm_perf_platform.tasks.scheduler import task_scheduler

router = APIRouter(prefix="/tests")

task_service = TaskService()
result_storage = ResultStorage()


@router.post("/run", response_model=TestRunResponse)
def run_test(request: TestRunRequest, current_user = Depends(get_current_user)):
    # 验证远程执行模式下必须提供SSH配置
    if request.execution_mode == "remote" and not request.ssh_config:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="SSH configuration is required for remote execution mode. Please configure SSH connection or switch to local mode."
        )

    parameters = request.model_dump()

    ssh_config_dict = None
    if request.ssh_config:
        ssh_config_dict = request.ssh_config.model_dump()

    record = task_service.create_task(
        engine=request.engine,
        model=request.model,
        parameters=parameters,
        status="queued",
        ssh_config=ssh_config_dict,
        user_id=current_user.id,
        appauto_branch=request.appauto_branch,
        task_type=parameters.get("task_type", "perf_test"),
    )

    # 构建调度器 payload - 使用 Pydantic 模型（旧版 API）
    payload = LegacyTestPayload(
        task_id=record.id,
        task_type="perf_test_api",
        engine=request.engine,
        model=request.model,
        input_length=request.input_length,
        output_length=request.output_length,
        concurrency=request.concurrency,
        loop=request.loop,
        warmup=request.warmup,
        execution_mode=request.execution_mode,
        scenario=request.scenario,
        ssh_config=ssh_config_dict,
        appauto_branch=request.appauto_branch,
        # 模型启动配置
        auto_launch_model=request.auto_launch_model,
        model_config_name=request.model_config_name,
        model_path=request.model_path,
        model_tp=request.model_tp,
        model_mode=request.model_mode,
        model_port=request.model_port,
        model_host=request.model_host,
        stop_model_after_test=request.stop_model_after_test,
        # AMaaS API 配置
        amaas_api_port=request.amaas_api_port,
        amaas_api_user=request.amaas_api_user,
        amaas_api_passwd=request.amaas_api_passwd,
    )

    task_scheduler.submit(record.id, payload)

    return TestRunResponse(
        task_id=record.id,
        status=record.status,
        concurrency=request.concurrency,
    )


@router.post("/run_perf/amaas/skip_launch", response_model=TestRunResponse)
def run_perf_via_amaas_skip_launch(request: TestPerfViaAMaaSSkipLaunch, current_user = Depends(get_current_user)):
    """AMaaS 场景性能测试 - 跳过模型启动"""
    return _run_perf_test(
        base="amaas",
        skip_launch=True,
        request=request,
        current_user=current_user,
    )


@router.post("/run_perf/amaas/with_launch", response_model=TestRunResponse)
def run_perf_via_amaas_with_launch(request: TestPerfViaAmaaSWithLaunch, current_user = Depends(get_current_user)):
    """AMaaS 场景性能测试 - 自动启动模型"""
    return _run_perf_test(
        base="amaas",
        skip_launch=False,
        request=request,
        current_user=current_user,
    )


@router.post("/run_perf/ft/skip_launch", response_model=TestRunResponse)
def run_perf_via_ft_skip_launch(request: TestPerfViaFTSkipLaunch, current_user = Depends(get_current_user)):
    """FT 容器场景性能测试 - 跳过模型启动"""
    return _run_perf_test(
        base="ft",
        skip_launch=True,
        request=request,
        current_user=current_user,
    )


@router.post("/run_perf/ft/with_launch", response_model=TestRunResponse)
def run_perf_via_ft_with_launch(request: TestPerfViaFTWithLaunch, current_user = Depends(get_current_user)):
    """FT 容器场景性能测试 - 自动启动模型"""
    return _run_perf_test(
        base="ft",
        skip_launch=False,
        request=request,
        current_user=current_user,
    )


@router.get("/list", response_model=TaskListResponse)
def list_tasks(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页大小"),
    task_type: str = Query(None, description="任务类型过滤 (hardware_info, perf_test, etc.)"),
    current_user = Depends(get_current_user)
):
    """获取任务列表（分页）

    默认每页返回20个任务，按创建时间倒序排列
    可以通过 task_type 参数过滤特定类型的任务
    """
    tasks, total = task_service.list_tasks_paginated(
        page=page,
        page_size=page_size,
        task_type=task_type,
    )

    items = [serialize_task(record, current_user.id) for record in tasks]
    total_pages = math.ceil(total / page_size) if total > 0 else 0

    return TaskListResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages
    )


@router.get("/{task_id}", response_model=TaskDetailResponse)
def get_task(task_id: int):
    record = task_service.get_task(task_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found")
    return serialize_task(record)


@router.post("/archive", response_model=ArchiveResponse)
def archive_task(request: ArchiveRequest):
    """归档任务结果文件

    ⚠️ WARNING: This API is under development and incomplete.
    TODO:
    - 当前仅归档 xlsx 文件，缺少硬件配置信息
    - 需要收集并打包硬件信息（GPU、CPU、内存等）
    - 需要添加测试配置参数
    - 归档包应包含：性能结果 + 硬件信息 + 测试配置

    使用建议：
    1. 先使用 /hardware_info/collect API 收集硬件信息
    2. 等待完整的归档功能实现后再使用本功能
    """
    record = task_service.get_task(request.task_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found")
    if not record.result_path:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Task result missing")

    archived_path = result_storage.archive_result(
        engine=record.engine,
        model=record.model,
        source_path=record.result_path,
    )
    task_service.attach_archive(record.id, archived_path)
    return ArchiveResponse(task_id=record.id, archived_path=archived_path)


@router.get("/{task_id}/result")
def download_result(task_id: int):
    record = task_service.get_task(task_id)
    if not record or not record.result_path:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Result not found")

    path = Path(record.result_path)
    if not path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File missing on disk")

    return FileResponse(path, filename=path.name)


@router.delete("/{task_id}", response_model=DeleteTaskResponse)
def delete_task(task_id: int, current_user = Depends(get_current_user)):
    record = task_service.get_task(task_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found")

    # 检查权限：只有任务创建者可以删除
    if record.user_id and record.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to delete this task"
        )

    task_service.delete_task(task_id)
    result_storage.delete_file(record.result_path)
    result_storage.delete_file(record.archived_path)

    # 删除日志文件 - 使用新的命名格式: {display_id}_{uuid}.log
    display_id = record.display_id or task_id
    log_file = LOG_DIR / f"{display_id}_{record.uuid}.log"
    if log_file.exists():
        log_file.unlink()
    return DeleteTaskResponse(deleted=True)


@router.post("/{task_id}/cancel", response_model=CancelTaskResponse)
def cancel_task(task_id: int, current_user = Depends(get_current_user)):
    """取消正在运行的任务

    只有任务创建者可以取消任务。
    只能取消状态为 queued 或 running 的任务。
    """
    # 获取任务记录
    record = task_service.get_task(task_id)
    if not record:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Task not found")

    # 检查权限：只有任务创建者可以取消
    if record.user_id and record.user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have permission to cancel this task"
        )

    # 检查任务状态
    if record.status not in ["queued", "running"]:
        return CancelTaskResponse(
            task_id=task_id,
            cancelled=False,
            message=f"Task cannot be cancelled. Current status: {record.status}"
        )

    # 尝试取消任务
    cancelled = task_scheduler.cancel_task(task_id)

    if cancelled:
        return CancelTaskResponse(
            task_id=task_id,
            cancelled=True,
            message="Task has been cancelled successfully"
        )
    else:
        # 如果 Future 无法取消（可能已经开始执行），直接标记为已取消
        task_service.mark_cancelled(task_id)
        return CancelTaskResponse(
            task_id=task_id,
            cancelled=True,
            message="Task has been marked as cancelled"
        )


@router.post("/{task_id}/retry", response_model=RetryTaskResponse)
def retry_task(task_id: int, current_user = Depends(get_current_user)):
    """重新提交任务

    根据原任务的参数创建一个新任务并提交执行。
    适用于失败或需要重新运行的任务。
    """
    # 获取原任务记录
    original_task = task_service.get_task(task_id)
    if not original_task:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Task {task_id} not found"
        )

    # 提取原任务参数
    parameters = dict(original_task.parameters)

    # 自动检测并添加缺失的 task_type 字段（向后兼容）
    if "task_type" not in parameters:
        # 通过 engine 字段判断任务类型
        if original_task.engine == "pytest":
            parameters["task_type"] = TaskType.PYTEST.value
        elif original_task.engine == "evalscope":
            parameters["task_type"] = TaskType.EVAL_TEST.value
        elif original_task.engine == "appauto":
            parameters["task_type"] = TaskType.ENV_DEPLOY.value
        # 如果有 base/skip_launch/parallel/number 字段，说明是命令行方式的 perf test
        elif "base" in parameters and "parallel" in parameters and "number" in parameters:
            parameters["task_type"] = "perf_test_cmd"

            # 命令行方式需要顶层的 SSH 字段，如果缺失则从 ssh_config 中提取
            ssh_config = parameters.get("ssh_config")
            if ssh_config:
                if "ssh_user" not in parameters:
                    parameters["ssh_user"] = ssh_config.get("user", "")
                if "ssh_password" not in parameters:
                    parameters["ssh_password"] = ssh_config.get("password")
                if "ssh_port" not in parameters:
                    parameters["ssh_port"] = ssh_config.get("port", 22)
        # 否则是 Python API 方式
        # 不需要显式设置，调度器会使用默认值 perf_test_api

    # 创建新任务记录
    new_record = task_service.create_task(
        engine=original_task.engine,
        model=original_task.model,
        parameters=parameters,
        status="queued",
        ssh_config=original_task.ssh_config,
        user_id=current_user.id,
        task_type=original_task.task_type,
        appauto_branch=parameters.get("appauto_branch", "main"),
    )

    # 构建 payload 并提交到调度器
    payload = dict(parameters)
    payload["task_id"] = new_record.id
    payload["engine"] = original_task.engine
    payload["model"] = original_task.model

    task_scheduler.submit(new_record.id, payload)

    return RetryTaskResponse(
        task_id=task_id,
        new_task_id=new_record.id,
        status="queued",
        message=f"Task {task_id} has been resubmitted as task {new_record.id}"
    )


@router.get("/{task_id}/logs", response_model=TaskLogsResponse)
def get_task_logs(task_id: int):
    """获取任务执行日志

    返回任务的完整执行日志内容。
    如果日志文件不存在，返回提示信息。
    """
    # 检查任务是否存在
    record = task_service.get_task(task_id)
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Task {task_id} not found"
        )

    # 获取日志文件路径 - 使用新的命名格式: {display_id}_{uuid}.log
    display_id = record.display_id or task_id
    log_file = LOG_DIR / f"{display_id}_{record.uuid}.log"

    if not log_file.exists():
        return TaskLogsResponse(
            task_id=task_id,
            logs=f"No logs available for task {task_id}. Task may not have started yet or logs were not generated.",
            log_file_path=None
        )

    # 读取日志内容
    try:
        with open(log_file, "r", encoding="utf-8") as f:
            logs = f.read()

        return TaskLogsResponse(
            task_id=task_id,
            logs=logs,
            log_file_path=str(log_file)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to read logs: {str(e)}"
        )


@router.post("/hardware_info/collect", response_model=HardwareInfoCollectResponse)
def collect_hardware_info(request: HardwareInfoCollectRequest, current_user = Depends(get_current_user)):
    """收集远程机器的硬件信息

    收集包括：
    - GPU 信息（nvidia-smi）
    - CPU 信息（核心数、型号、架构）
    - 内存信息（总量、已用、可用）
    - 磁盘信息
    - 操作系统信息
    - 网络信息

    生成的 JSON 文件可供下载查看。
    """
    # 构建 SSH 配置
    ssh_config = request.ssh_config.model_dump()

    # 创建任务记录
    record = task_service.create_task(
        engine="system",  # 使用 system 作为引擎标识
        model="hardware_info",  # 使用 hardware_info 作为模型名称
        parameters={
            "ssh_config": ssh_config,
            "timeout": request.timeout,
        },
        status="queued",
        ssh_config=ssh_config,
        user_id=current_user.id,
        task_type="hardware_info",
    )

    # 构建调度器 payload - 使用 Pydantic 模型
    payload = HardwareInfoPayload(
        task_id=record.id,
        task_type="hardware_info",
        ssh_config=ssh_config,
        timeout=request.timeout,
    )

    # 提交任务到调度器
    task_scheduler.submit(record.id, payload)

    return HardwareInfoCollectResponse(
        task_id=record.id,
        status=record.status,
        message=f"Hardware info collection task {record.id} has been submitted"
    )


def _run_perf_test(
    base: str,
    skip_launch: bool,
    request,
    current_user,
) -> TestRunResponse:
    """性能测试的通用实现函数

    Args:
        base: 测试基础环境，"amaas" 或 "ft"
        skip_launch: 是否跳过模型启动
        request: 请求对象（BaseTestPerf 子类）
        current_user: 当前登录用户

    Returns:
        TestRunResponse: 测试任务响应
    """
    # 构建 SSH 配置
    ssh_config = {
        "host": request.ip,
        "port": request.ssh_port,
        "user": request.ssh_user,
        "auth_type": "password" if request.ssh_password else "key",
        "password": request.ssh_password,
        "timeout": 30,
    }

    # 解析并发度参数
    parallel_list = [int(x.strip()) for x in request.parallel.split()]

    # 估算建议并发度（使用第一个并发值作为参考）
    concurrency_details: Dict[str, Any] | None = None
    suggested_concurrency = parallel_list[0] if parallel_list else 1

    # 准备模型名称（如果没有提供，则留空等待验证）
    model = request.model or "unknown"

    # 构建任务参数（包含所有 payload 需要的字段，确保重试时不丢失信息）
    parameters = {
        "task_type": "perf_test_cmd",  # 标记为命令行方式，重试时需要
        "base": base,
        "skip_launch": skip_launch,
        "ip": request.ip,
        "port": request.port,
        "model": model,
        "tokenizer_path": request.tokenizer_path,
        "ssh_config": ssh_config,
        "ssh_user": request.ssh_user,
        "ssh_password": request.ssh_password,
        "ssh_port": request.ssh_port,
        "parallel": request.parallel,
        "number": request.number,
        "input_length": request.input_length,
        "output_length": request.output_length,
        "loop": request.loop,
        "debug": request.debug,
        "warmup": request.warmup,
        "keep_model": request.keep_model,
        "concurrency": suggested_concurrency,
        "tp": request.tp,
        "appauto_branch": request.appauto_branch,
        "timeout_minutes": request.timeout_minutes,
    }

    # 根据场景添加特定参数
    if not skip_launch:
        if base == "amaas":
            parameters["amaas_api_port"] = 10001
            parameters["amaas_api_user"] = "admin"
            parameters["amaas_api_passwd"] = "123456"
        elif base == "ft":
            parameters["launch_timeout"] = getattr(request, "launch_timeout", 900)

    # 创建任务记录
    record = task_service.create_task(
        engine="evalscope",  # 使用 evalscope 作为引擎
        model=model,
        parameters=parameters,
        status="queued",
        ssh_config=ssh_config,
        user_id=current_user.id,
        appauto_branch=request.appauto_branch,
        task_type="perf_test",
    )

    # 构建调度器 payload - 使用 Pydantic 模型
    payload_dict = {
        "task_id": record.id,
        "task_type": "perf_test_cmd",
        "base": base,
        "skip_launch": skip_launch,
        "ip": request.ip,
        "port": request.port,
        "model": model,
        "tokenizer_path": request.tokenizer_path,
        "ssh_user": request.ssh_user,
        "ssh_password": request.ssh_password,
        "ssh_port": request.ssh_port,
        "parallel": request.parallel,
        "number": request.number,
        "input_length": request.input_length,
        "output_length": request.output_length,
        "loop": request.loop,
        "debug": request.debug,
        "warmup": request.warmup,
        "keep_model": request.keep_model,
        "ssh_config": ssh_config,
        "tp": request.tp,
        "appauto_branch": request.appauto_branch,
        "timeout_minutes": request.timeout_minutes,
    }

    # 添加场景特定参数
    if not skip_launch:
        if base == "amaas":
            payload_dict["amaas_api_port"] = 10001
            payload_dict["amaas_api_user"] = "admin"
            payload_dict["amaas_api_passwd"] = "123456"
        elif base == "ft":
            payload_dict["tp"] = getattr(request, "tp", 1)
            payload_dict["launch_timeout"] = getattr(request, "launch_timeout", 900)

    # 创建 Pydantic Payload 模型（会自动验证）
    payload = PerfTestPayload(**payload_dict)

    # 提交任务到调度器
    task_scheduler.submit(record.id, payload)

    return TestRunResponse(
        task_id=record.id,
        status=record.status,
        concurrency=suggested_concurrency,
        auto_concurrency=False,
        concurrency_details=concurrency_details,
    )


def serialize_task(record, current_user_id: int = None) -> TaskSummary:
    user_email = None
    if record.user_id:
        with get_session() as session:
            user = session.get(UserAccount, record.user_id)
            if user:
                user_email = user.email

    return TaskSummary(
        id=record.id,
        uuid=record.uuid,
        display_id=record.display_id,
        engine=record.engine,
        model=record.model,
        status=record.status,
        created_at=record.created_at,
        completed_at=record.completed_at,
        result_path=record.result_path,
        archived_path=record.archived_path,
        error_message=record.error_message,
        parameters=record.parameters,
        summary=record.summary,
        ssh_config=record.ssh_config,
        user_id=record.user_id,
        user_email=user_email,
    )


@router.get("/appauto/branches")
def get_appauto_branches(current_user: UserAccount = Depends(get_current_user)) -> Dict[str, Any]:
    """获取可用的 appauto 分支列表

    从 appauto 源码仓库获取所有分支
    """
    import subprocess
    from llm_perf_platform.services.venv_manager import get_venv_manager

    venv_manager = get_venv_manager()

    # 新架构：使用任意一个现有的仓库来查询分支信息
    # 所有仓库都指向同一个远程，所以任何一个都可以
    appauto_repo = venv_manager.get_any_repo_path()

    if not appauto_repo:
        # 如果还没有任何仓库，先创建一个 main 分支的仓库
        logger.info("No existing repositories found, creating main branch repository")
        if not venv_manager.create_repo("main"):
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to create appauto repository. Please check git configuration."
            )
        appauto_repo = venv_manager.get_repo_path("main")

    if not appauto_repo.exists():
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Appauto repository not found at {appauto_repo}"
        )

    try:
        # 获取所有本地和远程分支
        # 设置 GIT_PAGER=cat 禁用交互式分页器
        env = os.environ.copy()
        env["GIT_PAGER"] = "cat"

        result = subprocess.run(
            ["git", "branch", "-a"],
            cwd=appauto_repo,
            capture_output=True,
            text=True,
            timeout=10,
            env=env,
        )

        if result.returncode != 0:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to get branches: {result.stderr}"
            )

        # 解析分支列表
        branches = []
        for line in result.stdout.split("\n"):
            line = line.strip()
            if not line:
                continue

            # 移除当前分支标记 *
            if line.startswith("* "):
                line = line[2:]

            # 跳过 HEAD 指针
            if "HEAD ->" in line:
                continue

            # 处理远程分支
            if line.startswith("remotes/origin/"):
                branch = line.replace("remotes/origin/", "")
                if branch not in branches:
                    branches.append(branch)
            else:
                # 本地分支
                if line not in branches:
                    branches.append(line)

        # 排序并将 main 放在最前面
        branches.sort()
        if "main" in branches:
            branches.remove("main")
            branches.insert(0, "main")

        return {
            "branches": branches,
            "source_path": str(appauto_repo),
        }

    except subprocess.TimeoutExpired:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Timeout while getting branches"
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error getting branches: {str(e)}"
        )


def extract_chart_data(all_rows: List[tuple]) -> Dict[str, Any] | None:
    """从Excel横向表格中提取图表数据

    Args:
        all_rows: Excel的所有行数据

    Returns:
        图表数据字典，包含:
        - concurrency_levels: 并发度列表
        - ttft_values: 各并发度的平均TTFT
        - tps_values: 各并发度的平均TPS
        - has_multiple_concurrency: 是否有多个并发度
        如果无法提取则返回None
    """
    if len(all_rows) < 2:
        return None

    # 第一行是表头
    headers = all_rows[0]
    if not headers:
        return None

    # 查找需要的列索引
    try:
        concurrency_idx = None
        ttft_idx = None
        tps_idx = None

        for i, header in enumerate(headers):
            if header == "Concurrency":
                concurrency_idx = i
            elif header == "Average time to first token (s)":
                ttft_idx = i
            elif header == "单轮次 Avg TPS":
                tps_idx = i

        # 检查是否找到所有必需的列
        if concurrency_idx is None or ttft_idx is None or tps_idx is None:
            return None

        # 按并发度分组收集数据
        concurrency_data = {}  # {concurrency: {"ttft": [...], "tps": [...]}}

        for row in all_rows[1:]:  # 跳过表头
            if len(row) <= max(concurrency_idx, ttft_idx, tps_idx):
                continue

            try:
                concurrency = int(row[concurrency_idx]) if row[concurrency_idx] else None
                ttft = float(row[ttft_idx]) if row[ttft_idx] is not None and row[ttft_idx] != "" else None
                tps = float(row[tps_idx]) if row[tps_idx] is not None and row[tps_idx] != "" else None

                if concurrency is not None and ttft is not None and tps is not None:
                    if concurrency not in concurrency_data:
                        concurrency_data[concurrency] = {"ttft": [], "tps": []}

                    concurrency_data[concurrency]["ttft"].append(ttft)
                    concurrency_data[concurrency]["tps"].append(tps)
            except (ValueError, TypeError):
                continue

        # 如果没有有效数据
        if not concurrency_data:
            return None

        # 计算每个并发度的平均值
        concurrency_levels = sorted(concurrency_data.keys())
        ttft_values = []
        tps_values = []

        for concurrency in concurrency_levels:
            ttft_list = concurrency_data[concurrency]["ttft"]
            tps_list = concurrency_data[concurrency]["tps"]

            # 计算平均值
            avg_ttft = sum(ttft_list) / len(ttft_list) if ttft_list else 0
            avg_tps = sum(tps_list) / len(tps_list) if tps_list else 0

            ttft_values.append(round(avg_ttft, 4))
            tps_values.append(round(avg_tps, 2))

        return {
            "concurrency_levels": concurrency_levels,
            "ttft_values": ttft_values,
            "tps_values": tps_values,
            "has_multiple_concurrency": len(concurrency_levels) > 1
        }

    except Exception:
        return None


@router.get("/{task_id}/preview")
def preview_result(task_id: int, current_user = Depends(get_current_user)):
    """预览测试结果的Excel数据

    返回JSON格式的数据，包含：
    - parameters: 测试参数
    - summary: 汇总指标
    - requests: 请求详情（可选，用于详细分析）
    """
    # 获取任务记录
    record = task_service.get_task(task_id)
    if not record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"任务 {task_id} 不存在"
        )

    # 检查结果文件是否存在
    if not record.result_path:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"任务 {task_id} 没有生成结果文件，可能尚未完成或执行失败"
        )

    result_path = Path(record.result_path)
    if not result_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"结果文件不存在：{record.result_path}（文件可能已被删除或移动）"
        )

    # 检测文件类型
    file_ext = result_path.suffix.lower()

    # 如果是 JSON 文件
    if file_ext == '.json':
        try:
            import json
            with open(result_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            return {
                "task_id": task_id,
                "file_type": "json",
                "json_data": json_data,
            }
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"解析 JSON 文件失败：{str(e)}"
            )

    try:
        # 解析Excel文件 - 通用方式，直接展示所有工作表的原始内容
        wb = load_workbook(result_path, read_only=True, data_only=True)

        sheets = []
        chart_data = None  # 图表数据
        max_preview_rows = 100  # 每个工作表最多预览100行

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            all_rows = []  # 用于图表数据提取，读取所有行

            # 读取工作表数据
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                row_data = [cell if cell is not None else "" for cell in row]
                all_rows.append(row)  # 保存原始数据用于图表

                if i < max_preview_rows:
                    rows.append(row_data)

            if rows:  # 只添加非空工作表
                sheets.append({
                    "name": sheet_name,
                    "rows": rows,
                    "total_rows": ws.max_row,
                    "is_truncated": ws.max_row > max_preview_rows
                })

            # 尝试提取图表数据（仅处理旧格式的横向表格）
            if sheet_name == "Sheet" and len(all_rows) >= 2:
                chart_data = extract_chart_data(all_rows)

        wb.close()

        return {
            "task_id": task_id,
            "sheets": sheets,
            "chart_data": chart_data,
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"解析 Excel 文件失败：{str(e)}"
        )


# ========== 正确性测试（EvalScope Eval）API ==========

def _run_eval_test(
    base: str,
    skip_launch: bool,
    request,
    current_user,
) -> TestEvalResponse:
    """正确性测试的通用实现函数
    
    Args:
        base: 测试基础环境，"amaas" 或 "ft"
        skip_launch: 是否跳过模型启动
        request: 请求对象（BaseTestEval 子类）
        current_user: 当前登录用户
    
    Returns:
        TestEvalResponse: 测试任务响应
    """
    # 构建 SSH 配置
    ssh_config = {
        "host": request.ip,
        "port": request.ssh_port,
        "user": request.ssh_user,
        "auth_type": "password" if request.ssh_password else "key",
        "password": request.ssh_password,
        "timeout": 30,
    }
    
    # 准备模型名称
    model = request.model
    
    # 构建任务参数（包含所有 payload 需要的字段，确保重试时不丢失信息）
    parameters = {
        "task_type": "eval_test",  # 标记为正确性测试
        "base": base,
        "skip_launch": skip_launch,
        "ip": request.ip,
        "port": request.port,
        "model": model,
        "ssh_config": ssh_config,
        "ssh_user": request.ssh_user,
        "ssh_password": request.ssh_password,
        "ssh_port": request.ssh_port,

        # 评测参数
        "dataset": request.dataset,
        "dataset_args": request.dataset_args,
        "max_tokens": request.max_tokens,
        "concurrency": request.concurrency,
        "limit": request.limit,
        "temperature": request.temperature,
        "enable_thinking": request.enable_thinking,
        "debug": request.debug,

        # 模型启动参数
        "tp": request.tp,
        "keep_model": request.keep_model,
        "appauto_branch": request.appauto_branch,
        "timeout": request.timeout,
    }
    
    # 根据场景添加特定参数
    if not skip_launch:
        parameters["launch_timeout"] = getattr(request, "launch_timeout", 900)
    
    # 创建任务记录
    record = task_service.create_task(
        engine="evalscope",  # 使用 evalscope 作为引擎
        model=model,
        parameters=parameters,
        status="queued",
        ssh_config=ssh_config,
        user_id=current_user.id,
        appauto_branch=request.appauto_branch,
        task_type="eval_test",
    )
    
    # 构建调度器 payload - 使用 Pydantic 模型
    payload_dict = {
        "task_id": record.id,
        "task_type": "eval_test",
        "base": base,
        "skip_launch": skip_launch,
        "ip": request.ip,
        "port": request.port,
        "model": model,
        "ssh_user": request.ssh_user,
        "ssh_password": request.ssh_password,
        "ssh_port": request.ssh_port,
        "ssh_config": ssh_config,

        # 评测参数
        "dataset": request.dataset,
        "dataset_args": request.dataset_args,
        "max_tokens": request.max_tokens,
        "concurrency": request.concurrency,
        "limit": request.limit,
        "temperature": request.temperature,
        "enable_thinking": request.enable_thinking,
        "debug": request.debug,

        # 模型启动参数
        "tp": request.tp,
        "keep_model": request.keep_model,
        "appauto_branch": request.appauto_branch,
        "timeout": request.timeout,
    }

    # 添加场景特定参数
    if not skip_launch:
        payload_dict["launch_timeout"] = getattr(request, "launch_timeout", 900)

    # 创建 Pydantic Payload 模型（会自动验证）
    payload = EvalTestPayload(**payload_dict)

    # 提交任务到调度器
    task_scheduler.submit(record.id, payload)
    
    return TestEvalResponse(
        task_id=record.id,
        display_id=record.display_id or record.id,
        uuid=record.uuid,
        status=record.status,
        message=f"正确性测试任务已创建: {record.display_id or record.id}",
    )


@router.post("/run_eval/amaas/skip_launch", response_model=TestEvalResponse)
def run_eval_via_amaas_skip_launch(request: TestEvalViaAMaaSSkipLaunch, current_user = Depends(get_current_user)):
    """AMaaS 场景正确性测试 - 跳过模型启动"""
    return _run_eval_test(
        base="amaas",
        skip_launch=True,
        request=request,
        current_user=current_user,
    )


@router.post("/run_eval/amaas/with_launch", response_model=TestEvalResponse)
def run_eval_via_amaas_with_launch(request: TestEvalViaAMaaSWithLaunch, current_user = Depends(get_current_user)):
    """AMaaS 场景正确性测试 - 自动启动模型"""
    return _run_eval_test(
        base="amaas",
        skip_launch=False,
        request=request,
        current_user=current_user,
    )


@router.post("/run_eval/ft/skip_launch", response_model=TestEvalResponse)
def run_eval_via_ft_skip_launch(request: TestEvalViaFTSkipLaunch, current_user = Depends(get_current_user)):
    """FT 容器场景正确性测试 - 跳过模型启动"""
    return _run_eval_test(
        base="ft",
        skip_launch=True,
        request=request,
        current_user=current_user,
    )


@router.post("/run_eval/ft/with_launch", response_model=TestEvalResponse)
def run_eval_via_ft_with_launch(request: TestEvalViaFTWithLaunch, current_user = Depends(get_current_user)):
    """FT 容器场景正确性测试 - 自动启动模型"""
    return _run_eval_test(
        base="ft",
        skip_launch=False,
        request=request,
        current_user=current_user,
    )


# ========== 环境部署 API ==========

def _run_deploy(
    deploy_type: str,
    request,
    current_user,
) -> DeployResponse:
    """环境部署的通用实现函数

    Args:
        deploy_type: 部署类型，"amaas" 或 "ft"
        request: 请求对象（DeployAMaaSRequest 或 DeployFTRequest）
        current_user: 当前登录用户

    Returns:
        DeployResponse: 部署任务响应
    """
    # 构建任务参数
    parameters = {
        "task_type": "env_deploy",  # 标记为环境部署
        "deploy_type": deploy_type,
        "ip": request.ip,
        "tar_name": request.tar_name,
        "ssh_user": request.ssh_user,
        "ssh_password": request.ssh_password,
        "ssh_port": request.ssh_port,
        "user": request.user,
        "appauto_branch": request.appauto_branch,
    }

    # AMaaS 部署需要 tag 参数（FT 已移除 tag）
    if deploy_type == "amaas":
        parameters["tag"] = request.tag

    # 创建任务记录
    record = task_service.create_task(
        engine="appauto",  # 使用 appauto 作为引擎
        model=f"{deploy_type}_deploy",  # 模型字段用来区分部署类型
        parameters=parameters,
        status="queued",
        ssh_config=None,  # 部署任务不需要 SSH 配置（参数已包含在 parameters 中）
        user_id=current_user.id,
        appauto_branch=request.appauto_branch,
        task_type="env_deploy",
    )

    # 构建调度器 payload - 使用 Pydantic 模型
    payload_dict = {
        "task_id": record.id,
        "task_type": "env_deploy",
        "deploy_type": deploy_type,
        "ip": request.ip,
        "tar_name": request.tar_name,
        "ssh_user": request.ssh_user,
        "ssh_password": request.ssh_password,
        "ssh_port": request.ssh_port,
        "user": request.user,
        "appauto_branch": request.appauto_branch,
    }

    # AMaaS 部署需要 tag 参数（FT 已移除 tag）
    if deploy_type == "amaas":
        payload_dict["tag"] = request.tag

    # 创建 Pydantic Payload 模型（会自动验证）
    payload = EnvDeployPayload(**payload_dict)

    # 提交任务到调度器
    task_scheduler.submit(record.id, payload)

    return DeployResponse(
        task_id=record.id,
        display_id=record.display_id or record.id,
        uuid=record.uuid,
        status=record.status,
        message=f"{deploy_type.upper()} 部署任务已创建: {record.display_id or record.id}",
    )


@router.post("/deploy/amaas", response_model=DeployResponse)
def deploy_amaas(request: DeployAMaaSRequest, current_user = Depends(get_current_user)):
    """部署 AMaaS 环境"""
    return _run_deploy(
        deploy_type="amaas",
        request=request,
        current_user=current_user,
    )


@router.post("/deploy/ft", response_model=DeployResponse)
def deploy_ft(request: DeployFTRequest, current_user = Depends(get_current_user)):
    """部署 FT 环境"""
    return _run_deploy(
        deploy_type="ft",
        request=request,
        current_user=current_user,
    )
